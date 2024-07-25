import os
from openpyxl import Workbook, load_workbook


# Function to split data into multiple Excel files
def split_excel(input_file, output_dir, chunk_size=10000):
    # Load the input workbook and sheet
    book = load_workbook(input_file)
    sheet = book['Sheet']

    # Create the output directory if it does not exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Initialize variables for chunking
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]  # Exclude the headers from data
    num_chunks = len(data) // chunk_size + (1 if len(data) % chunk_size else 0)

    # Split the data and write to new files
    for i in range(num_chunks):
        chunk_data = data[i * chunk_size:(i + 1) * chunk_size]
        output_book = Workbook()
        output_sheet = output_book.active

        # Write headers to the new sheet only if it's the last chunk
        if i == num_chunks - 1:
            output_sheet.append(headers)

        # Write chunk data to the new sheet
        for row in chunk_data:
            output_sheet.append(row)

        # Save the new workbook
        output_file = os.path.join(output_dir, f'SafPromo25th{i + 1}.xlsx')
        output_book.save(output_file)
        print(f'Saved {output_file}')


# Example usage
input_file = "/Users/bonnymonah/Desktop/AWS Cloud SAA03/Important/SD_Dispatch/Thursday 25th July 2024-200K MSISDN's- List 2.xlsx"
output_dir = "/Users/bonnymonah/Desktop/AWS Cloud SAA03/Important/SD_Dispatch"
split_excel(input_file, output_dir)
